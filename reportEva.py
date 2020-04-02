# -*- coding: utf-8 -*-
# Берем из монго, фильтруем по PRODUCTS, подставляем статусы и агентов из postgres, результат в Excel

import sys, argparse
from _datetime import datetime, timedelta, date
import os
from mysql.connector import MySQLConnection, Error
from collections import OrderedDict
import openpyxl
from pymongo import MongoClient
import psycopg2

from lib import read_config, l

#PRODUCTS = ['raiffeisen_loan_referral', 'raiffeisen_loan_lead_referral', 'raiffeisen_credit_card_referral',
#            'raiffeisen_credit_card_lead_referral']
#PRODUCTS = ['alfabank_100_v2']
PRODUCTS = ['openbank_debit_card_referral','openbank_loan_referral','openbank_refinancing_loan_referral',
            'openbank_refinancing_mortgage_referral','openbank_credit_card_referral']
#PRODUCTS = ['rosfin_poll']
BAD_FIELDS = ['_id','owner_id','client','callcenter_status_code', 'question_list']

STATUSES = {0: 'Новая заявка', 20: 'Новая заявка', 100: 'Заявка отправлена в очередь', 110: 'Введен СМС код',
            120: 'Запрошена повторная СМС', 130: 'В процессе', 140: 'Одобрена', 150: 'Предварительно одобрена',
            200: 'Завершено успешно', 400: 'Удалена', 410: 'Неизвестный статус', 430: 'Отказ',
            10: 'Анкета отредактирована', 500: 'Отладка', 510: 'Отложена', 420: 'Ошибка выгрузки',
            470: 'Ошибка в заявке', 600: 'Ожидает оплаты', 610: 'Оплачено', 620: 'Услуга получена',
            210: 'Займ выдан', 220: 'Займ выдан повторно', 230: 'Займ выдан через call-центр', 160: 'Заявка заполнена',
            170: 'Анкета успешно отправлена', 180: 'Ошибка отправки файлов', 190: 'Файлы успешно отправлены',
            50: 'Ошибка', 310: 'Заявка отправлена', 320: 'Сканы отправлены', 330: 'Документы отправлены',
            340: 'Ошибка', 350: 'Смс отправлена', 360: 'Завершено', 370: 'ЕСИА код отправлен', 440: 'Отказ клиента',
            450: 'Закрыта', 460: 'Истек срок действия решения Банка', 240: 'Приложение установлено',
            250: 'Счет пополнен', 260: 'Карта активирована', 1100: 'Заявка создана', 1200: 'Пройден прескоринг',
            1210: 'Не пройден прескоринг', 1300: 'Пройден скоринг', 1310: 'Не пройден скоринг', 1500: 'Карта выдана',
            1600: 'Карта активирована', 700: 'Карта не выдана', 710: 'Карта выдана', 750: 'Карта выдана',
            760: 'Карта активирована', 800: 'Анкета успешно заполнена', 900: 'Одобрено', 910: 'Отклонено',
            920: 'В ожидании', 990: 'Заявка передана', 2000: 'Принято', 2100: 'В обработке', 2200: 'Отклонено',
            850: 'В обработке', 860: 'Дубль заявки', 870: 'Данные не валидны', 880: 'Заявка передана'}

QUESTIONS = ['financial_state','financial_strategy','savings_strategy','savings_state','savings_target',
             'savings_method','savings_insurance','personal_credit','personal_credit_debt','personal_accounting',
             'savings_safest_method','savings_profitable_method','product_analytics','mlm_awareness','insurance_state',
             'pension_awareness','pension_contract','pension_payments_awareness','information_reliable_source',
             'secured_rights','secured_rights_police','financial_education_level','financial_education_sufficient',
             'financial_education_update','education_conference','education_conference_theme',
             'information_source_list','financial_subject_school']

WORK_FIELDS = {'organization': 'ООО "РИЧРИНГ"', 'organization_inn': '3025008374',
               'position': 'Ассистент коммерческого директора', 'salary': 70000, 'contact_phone': '74951201145',
               'additional_phone': '89894416401'}



agents = {}
#q1 = """
# Загружаем агентов из postgres
cfg = read_config(filename='anketa.ini', section='postgresql')
conn = psycopg2.connect(**cfg)
cursor = conn.cursor()
cursor.execute('select ac.id, ac.lastname, ac.name, ac.middlename, di.title from account as ac '
               'left join division as di on ac.division_id = di.id;')
for row in cursor:
    family = ''
    name = ''
    lastname = ''
    if row[1]:
        family = row[1]
    if row[2]:
        name = ' ' + row[2]
    if row[3]:
        lastname = ' ' + row[3]
    agents[row[0]] = [family + name + lastname, row[4]]
#"""

# подключаемся к серверу
cfg = read_config(filename='anketa.ini', section='Mongo')
conn = MongoClient('mongodb://' + cfg['user'] + ':' + cfg['password'] + '@' + cfg['ip'] + ':' + cfg['port'] + '/'
                   + cfg['db'])
# выбираем базу данных
db = conn.saturn_v

# выбираем коллекцию документов
colls = db.Products

#product_alias: raiffeisen_loan_referral
#product_alias: raiffeisen_loan_lead_referral

wb = openpyxl.load_workbook(filename='key.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]
name_of_categories = []
for i, row in enumerate(ws):
    if i > 0:
        break
    for j, cell in enumerate(row):
        if j < 4:
            continue
        name_of_categories.append(cell.value)

categories = {}
name_of_questions = {}
for i, row in enumerate(ws):
    if i < 1:
        continue
    for j, cell in enumerate(row):
        if j == 0:
            question = QUESTIONS[int(cell.value) - 1]
        elif j == 1:
            name_of_question = cell.value
        elif j == 2:
            n_answer = 100 * int(cell.value)
        elif j == 3:
            answer = cell.value
        elif j > 3:
            if name_of_questions.get(question, None):
                if not name_of_questions[question]:
                    name_of_questions[question] = name_of_question
            else:
                name_of_questions[question] = name_of_question
            if categories.get(question, None):
                categories[question][n_answer] = answer
            else:
                categories[question] = {n_answer: answer}
            continue

# Реверс
question_of_names = {}
for name_of_question in name_of_questions:
    question_of_names[name_of_questions[name_of_question]] = name_of_question

wb_rez = openpyxl.Workbook(write_only=True)
ws_rez = []

for i, product in enumerate(PRODUCTS):
    ws_rez.append(wb_rez.create_sheet(product))
    xlsx_lines = []
    fields = []
    first_anketa = True
    first_work = True
    for j, coll in enumerate(colls.find({'product_alias': product})):
        #print('\n\ncoll\n', coll, '\n', agents[485])
        for field in coll.keys():
            if field not in BAD_FIELDS and field not in fields:
                fields.append(field)
        #print('\n\nЗАГОЛОВОК\n', ['ФИО Агента', 'Организация'] + fields)
        #print('\nBEFORE agents\n',  agents[coll['owner_id']], '\n', coll['owner_id'])
        if agents.get(coll['owner_id']):
            fields_rez = [agents[coll['owner_id']][0], agents[coll['owner_id']][1]]
            #print('\nagents\n', fields_rez, '\n', agents[coll['owner_id']], '\n', coll['owner_id'])
        else:
            fields_rez = [str(coll['owner_id']), str(coll['owner_id'])]
        fields_anketa = {}
        fields_work = {}
        # Если есть question_list то добавляем поля анкеты в общий список полей и формируем словарь значений
        if 'question_list' in coll.keys():
            if first_anketa:
                for quest in QUESTIONS:
                    fields.append(name_of_questions[quest])
                first_anketa = False
            aqs = coll.get('question_list')
            for quest in QUESTIONS:
                # Если в поле список - конвертируем
                if str(type(aqs[quest])).find('list') > -1:
                    modul_string = ''
                    for aq in aqs[quest]:
                        modul_string += categories[quest][aq] + ', '
                    fields_anketa[name_of_questions[quest]] = modul_string.strip(' ').strip(',')
                else:
                    fields_anketa[name_of_questions[quest]] = categories[quest].get(aqs[quest],str(aqs[quest]) +
                                                                                    'нет расшифровки в key.xlsx')
        if 'work' in coll.keys():
            if first_work:
                fields += WORK_FIELDS.keys()
                first_work = False
            aqs = coll.get('work')
            for quest in WORK_FIELDS.keys():
                if aqs.get(quest, None):
                    fields_work[quest] = aqs[quest]
        for field in fields:
            if field == 'state_code':
                fields_rez.append(STATUSES[l(coll.get(field))])
            elif field in name_of_questions.values():
                fields_rez.append(fields_anketa[field])
            elif field in WORK_FIELDS.keys():
                if fields_work.get(field, None):
                    fields_rez.append(fields_work[field])
            else:
                if str(type(coll.get(field))).find('str') < 0 and str(type(coll.get(field))).find('int') < 0:
                    fields_rez.append(str(coll.get(field)))
                else:
                    fields_rez.append(coll.get(field))
        #print('\n\nfields_rez\n',fields_rez)
        xlsx_lines.append(fields_rez)
    ws_rez[i].append(['ФИО Агента', 'Организация'] + fields)
    for xlsx_line in xlsx_lines:
        ws_rez[i].append(xlsx_line)
wb_rez.save(datetime.now().strftime('%Y-%m-%d_%H-%M') + 'отчеты.xlsx')
